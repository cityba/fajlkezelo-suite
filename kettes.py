import os
import re
import sys
import subprocess
import time
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTreeWidget,
    QTreeWidgetItem, QFileDialog, QAbstractItemView, QHeaderView,
    QLabel, QCheckBox, QMessageBox, QProgressBar, QLineEdit,
    QFrame, QApplication, QGroupBox
)
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from docx import Document
from openpyxl import Workbook
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from PyQt5.QtGui import QPalette, QColor

def read_file_content(file_path):
    try:
        file_size = os.path.getsize(file_path)
        if file_size > 10 * 1024 * 1024:
            return None
        
        if file_path.endswith('.docx'):
            doc = Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        elif file_path.endswith('.xlsx'):
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            content = ""
            for row in sheet.iter_rows(values_only=True):
                content += " ".join(str(cell) for cell in row if cell) + "\n"
            return content
        elif file_path.endswith('.pdf'):
            reader = PdfReader(file_path)
            content = ""
            for page in reader.pages:
                content += page.extract_text() + "\n"
            return content
        else:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                return file.read()
    except Exception as e:
        print(f"Hiba történt a fájl olvasása közben: {e}")
        return None

def open_with_application(file_path):
    normalized_path = os.path.normpath(file_path)
    
    if os.path.isdir(normalized_path):
        if os.name == 'nt':
            subprocess.Popen(f'explorer "{normalized_path}"', shell=True)
        elif os.name == 'posix':
            subprocess.Popen(["xdg-open", normalized_path])
    elif normalized_path.endswith(('.zip', '.rar')) and os.name == 'nt':
        subprocess.Popen(f'start winrar "{normalized_path}"', shell=True)
    else:
        if os.name == 'nt':
            subprocess.Popen(f'start "" "{normalized_path}"', shell=True)
        elif os.name == 'posix':
            subprocess.Popen(["xdg-open", normalized_path])

class SearchWorker(QThread):
    update_progress = pyqtSignal(int, int, int, float)
    file_found_single = pyqtSignal(str, int)
    search_finished = pyqtSignal()
    status_update = pyqtSignal(str)

    def __init__(self, folder, pattern, exact_match, exclude_extensions, 
                 search_filenames_only, search_folders_only, start_date, end_date):
        super().__init__()
        self.folder = folder
        self.pattern = pattern
        self.exact_match = exact_match
        self.exclude_extensions = exclude_extensions
        self.search_filenames_only = search_filenames_only
        self.search_folders_only = search_folders_only
        self.start_date = start_date
        self.end_date = end_date
        self.stop_flag = False
        self.compiled_pattern = re.compile(pattern, re.IGNORECASE if not exact_match else 0)
        
        self.excluded_extensions = [
            ".exe", ".mp3", ".mp4", ".wav", ".jpg", ".jpeg", ".png", ".gif", ".bmp",
            ".tiff", ".psd", ".avi", ".mov", ".mkv", ".flv", ".wmv", ".flac", ".aac",
            ".ogg", ".wma", ".zip", ".rar", ".7z", ".tar", ".gz", ".bin", ".iso", ".dll",
            ".so", ".ttf", ".otf", ".woff", ".cbr", ".cbz", ".epub", ".mobi", ".db", ".sqlite",
            ".mdb", ".sys", ".msi", ".cab"
        ]

    def run(self):
        try:
            self.status_update.emit("Fájlok listázása...")
            file_list = []
            total_files = 0
            
            for entry in os.scandir(self.folder):
                if self.stop_flag:
                    return
                    
                if entry.is_dir(follow_symlinks=False):
                    for root, dirs, files in os.walk(entry.path):
                        if self.stop_flag:
                            return
                        if self.search_folders_only:
                            items = dirs
                        else:
                            items = files
                        for item in items:
                            file_list.append(os.path.join(root, item))
                            total_files += 1
                else:
                    file_list.append(entry.path)
                    total_files += 1
            
            if total_files == 0:
                self.search_finished.emit()
                return
                
            processed_files = 0
            found_files = 0
            start_time = time.time()
            
            num_workers = max(1, os.cpu_count() - 1)
            batch_size = min(100, max(10, total_files // 100))
            
            def process_file(file_path):
                if self.stop_flag:
                    return (file_path, 0)
                
                item_name = os.path.basename(file_path)
                match_count = 0
                
                if self.start_date or self.end_date:
                    try:
                        creation_time = os.path.getctime(file_path)
                        creation_date = datetime.fromtimestamp(creation_time).date()
                        if self.start_date and creation_date < self.start_date:
                            return (file_path, 0)
                        if self.end_date and creation_date > self.end_date:
                            return (file_path, 0)
                    except:
                        pass
                
                if self.exclude_extensions and not self.search_folders_only:
                    if any(item_name.lower().endswith(ext) for ext in self.excluded_extensions):
                        return (file_path, 0)
                
                if self.search_filenames_only or self.search_folders_only:
                    if self.compiled_pattern.search(item_name):
                        match_count = 1
                else:
                    try:
                        file_size = os.path.getsize(file_path)
                        if file_size > 10 * 1024 * 1024:
                            if self.compiled_pattern.search(item_name):
                                match_count = 1
                        else:
                            content = read_file_content(file_path)
                            if content:
                                match_count = len(self.compiled_pattern.findall(content))
                    except:
                        if self.compiled_pattern.search(item_name):
                            match_count = 1
                
                if match_count > 0:
                    self.file_found_single.emit(file_path, match_count)
                return (file_path, match_count)
            
            with ThreadPoolExecutor( max_workers=num_workers) as executor:
                futures = [executor.submit(process_file, fp) for fp in file_list]
                
                for future in as_completed(futures):
                    if self.stop_flag:
                        for f in futures:
                            f.cancel()
                        break
                        
                    file_path, match_count = future.result()
                    processed_files += 1
                    
                    if match_count > 0:
                        found_files += 1
                    
                    if processed_files % batch_size == 0 or processed_files == total_files:
                        elapsed = time.time() - start_time
                        self.update_progress.emit(processed_files, total_files, found_files, elapsed)
            
            self.status_update.emit("Keresés befejezve")
            self.search_finished.emit()
        except Exception as e:
            self.status_update.emit(f"Hiba: {str(e)}")

    def stop(self):
        self.stop_flag = True

class FileSearchApp(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.search_worker = None
        self.results = []
        self.set_dark_palette()
        self.init_ui()
        self.set_style()

    def set_dark_palette(self):
        app = QApplication.instance()
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(25, 25, 25))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, Qt.black)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, Qt.black)
        app.setPalette(palette)
        
    def set_style(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #353535;
                color: #FFFFFF;
                font-family: 'Segoe UI';
            }
            QLabel {
                color: #BBBBBB;
                font-size: 10pt;
            }
            QPushButton {
                background-color: #5799AD;
                color: white;
                padding: 5px 5px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 10pt;
                min-width: 100px;
                border: 1px solid #444444;
            }
            QPushButton:hover {
                background-color: #2C546B;
            }
            QPushButton:disabled {
                background-color: #555555;
                color: #AAAAAA;
            }
            QPushButton#danger {
                background-color: #E74C3C;
                border: none;
            }
            QPushButton#danger:hover {
                background-color: #C0392B;
            }
            QPushButton#success {
                background-color: #27AE60;
                border: none;
            }
            QPushButton#success:hover {
                background-color: #219653;
            }
            QLineEdit, QTreeWidget {
                background-color: #2D2D2D;
                border: 1px solid #555555;
                border-radius: 3px;
                padding: 5px;
                color: #DDDDDD;
                font-size: 10pt;
            }
            QTreeWidget {
                alternate-background-color: #353535;
            }
            QHeaderView::section {
                background-color: #444444;
                color: white;
                padding: 5px;
                font-weight: bold;
                border: 1px solid #555555;
            }
            QProgressBar {
                border: 1px solid #555555;
                border-radius: 3px;
                text-align: center;
                height: 20px;
                color: white;
            }
            QProgressBar::chunk {
                background-color: #27AE60;
                width: 10px;
            }
            QCheckBox {
                font-size: 10pt;
                color: #BBBBBB;
            }
            QGroupBox {
                border: 1px solid #555555;
                border-radius: 5px;
                margin-top: 1ex;
                font-weight: bold;
                color: #FFFFFF;
                background-color: #3A3A3A;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 5px;
            }
        """)

    def init_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        title = QLabel("Fájlba Kereső Program")
        title.setStyleSheet("font-size: 16pt; font-weight: bold; color: #FFFFFF;")
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        folder_layout = QHBoxLayout()
        lbl_folder = QLabel("Mappa kiválasztása:")
        self.folder_entry = QLineEdit()
        btn_browse = QPushButton("Tallózás")
        btn_browse.clicked.connect(self.browse_folder)
        
        folder_layout.addWidget(lbl_folder)
        folder_layout.addWidget(self.folder_entry, 1)
        folder_layout.addWidget(btn_browse)
        main_layout.addLayout(folder_layout)

        search_layout = QHBoxLayout()
        lbl_search = QLabel("Kereső kifejezés:")
        self.search_entry = QLineEdit()
        search_layout.addWidget(lbl_search)
        search_layout.addWidget(self.search_entry, 1)
        main_layout.addLayout(search_layout)

        options_frame = QGroupBox("Keresési beállítások")
        options_layout = QVBoxLayout(options_frame)
        
        self.exact_match = QCheckBox("Pontos egyezés")
        self.exclude_ext = QCheckBox("Fájlkiterjesztések kizárása")
        self.search_filenames = QCheckBox("Csak fájlnevekben keresés")
        self.search_folders = QCheckBox("Csak mappanevekben keresés")
        
        options_layout.addWidget(self.exact_match)
        options_layout.addWidget(self.exclude_ext)
        options_layout.addWidget(self.search_filenames)
        options_layout.addWidget(self.search_folders)
        
        date_layout = QHBoxLayout()
        lbl_date = QLabel("Létrehozás dátuma:")
        self.start_date = QLineEdit()
        self.start_date.setPlaceholderText("ÉÉÉÉ-HH-NN")
        self.end_date = QLineEdit()
        self.end_date.setPlaceholderText("ÉÉÉÉ-HH-NN")
        
        mai_datum = datetime.now().strftime("%Y-%m-%d")
        self.start_date.setText(mai_datum)
        self.end_date.setText(mai_datum)
        
        date_layout.addWidget(lbl_date)
        date_layout.addWidget(self.start_date)
        date_layout.addWidget(QLabel(" - "))
        date_layout.addWidget(self.end_date)
        date_layout.addStretch()
          
        options_layout.addLayout(date_layout)
        main_layout.addWidget(options_frame)

        button_layout = QHBoxLayout()
        self.btn_search = QPushButton("Keresés")
        self.btn_search.clicked.connect(self.start_search)
        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setObjectName("danger")
        self.btn_stop.clicked.connect(self.stop_search)
        self.btn_stop.setEnabled(False)
        self.btn_save = QPushButton("Lista mentése")
        self.btn_save.setObjectName("success")
        self.btn_save.clicked.connect(self.save_results)
        
        button_layout.addWidget(self.btn_search)
        button_layout.addWidget(self.btn_stop)
        button_layout.addWidget(self.btn_save)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)

        results_frame = QGroupBox("Eredmények")
        results_layout = QVBoxLayout(results_frame)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Fájl", "Találatok", "Műveletek"])
        self.tree.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tree.setSortingEnabled(True)
        self.tree.setAlternatingRowColors(True)
        self.tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tree.setMinimumHeight(300)
        
        results_layout.addWidget(self.tree)
        main_layout.addWidget(results_frame, 1)

        status_layout = QHBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        
        self.status_label = QLabel("Kész a keresésre")
        self.status_label.setStyleSheet("font-weight: bold; color: #CCCCCC;")
        
        status_layout.addWidget(self.progress_bar, 2)
        status_layout.addWidget(self.status_label, 1)
        main_layout.addLayout(status_layout)

        self.setLayout(main_layout)
        self.setWindowTitle("Fájlba Kereső Program")
        self.resize(800, 600)

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Válassz mappát")
        if folder:
            self.folder_entry.setText(folder)

    def start_search(self):
        folder = self.folder_entry.text()
        pattern = self.search_entry.text()
        
        if not folder or not pattern:
            QMessageBox.warning(self, "Hiányzó adat", "Kérlek válassz mappát és adj meg kereső kifejezést!")
            return
            
        start_date = None
        end_date = None
        
        if self.start_date.text():
            try:
                start_date = datetime.strptime(self.start_date.text(), "%Y-%m-%d").date()
            except:
                QMessageBox.warning(self, "Hibás dátum", "Érvénytelen kezdő dátum formátum! Használd az ÉÉÉÉ-HH-NN formátumot.")
                return
                
        if self.end_date.text():
            try:
                end_date = datetime.strptime(self.end_date.text(), "%Y-%m-%d").date()
            except:
                QMessageBox.warning(self, "Hibás dátum", "Érvénytelen záró dátum formátum! Használd az ÉÉÉÉ-HH-NN formátumot.")
                return
        
        self.tree.clear()
        self.results = []
        
        self.btn_search.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Keresés folyamatban...")
        
        self.search_worker = SearchWorker(
            folder,
            pattern,
            self.exact_match.isChecked(),
            self.exclude_ext.isChecked(),
            self.search_filenames.isChecked(),
            self.search_folders.isChecked(),
            start_date,
            end_date
        )
        
        self.search_worker.update_progress.connect(self.update_progress)
        self.search_worker.file_found_single.connect(self.add_result)
        self.search_worker.search_finished.connect(self.search_finished)
        self.search_worker.status_update.connect(self.status_label.setText)
        self.search_worker.start()

    def stop_search(self):
        if self.search_worker:
            self.search_worker.stop()
        self.btn_search.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.status_label.setText("Keresés leállítva")

    def update_progress(self, processed, total, found, elapsed_time):
        if total > 0:
            progress = int((processed / total) * 100)
            self.progress_bar.setValue(progress)
            
            if processed > 0:
                remaining = (elapsed_time / processed) * (total - processed)
                mins, secs = divmod(int(remaining), 60)
                time_str = f"{mins} perc {secs} mp"
            else:
                time_str = "számítás alatt"
                
            self.status_label.setText(
                f"Folyamat: {progress}% | "
                f"Feldolgozva: {processed}/{total} | "
                f"Találatok: {found} | "
                f"Hátralévő idő: {time_str}"
            )

    def add_result(self, file_path, match_count):
        self.results.append((file_path, match_count))
        
        item = QTreeWidgetItem([
            file_path,
            str(match_count)
        ])
        
        button_frame = QWidget()
        button_layout = QHBoxLayout(button_frame)
        button_layout.setContentsMargins(0, 0, 0, 0)
        
        btn_open = QPushButton("Fájl")
        btn_open.setFixedWidth(40)
        btn_open.clicked.connect(lambda checked, p=file_path: open_with_application(p))
        
        btn_folder = QPushButton("Mappa")
        btn_folder.setFixedWidth(40)
        btn_folder.clicked.connect(lambda checked, p=file_path: open_with_application(os.path.dirname(p)))
        
        button_layout.addWidget(btn_open)
        button_layout.addWidget(btn_folder)
        
        self.tree.addTopLevelItem(item)
        self.tree.setItemWidget(item, 2, button_frame)
        
        self.tree.scrollToItem(item)

    def search_finished(self):
        self.btn_search.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.status_label.setText(f"Keresés befejezve! Találatok: {len(self.results)}")

    def save_results(self):
        if not self.results:
            QMessageBox.information(self, "Nincs adat", "Nincs mentendő eredmény!")
            return
            
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Eredmények mentése",
            f"{self.search_entry.text() or 'eredmeny'}.xlsx",
            "Excel fájlok (*.xlsx)"
        )
        
        if not save_path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Eredmények"
            
            ws.append(["Fájl elérési út", "Találatok száma"])
            
            for file_path, count in self.results:
                ws.append([file_path, count])
            
            wb.save(save_path)
            QMessageBox.information(self, "Sikeres mentés", f"Eredmények mentve: {save_path}")
            
            self.show_save_buttons(save_path)
            
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Mentés sikertelen!\n{str(e)}")

    def show_save_buttons(self, save_path):
        for i in reversed(range(self.layout().count())):
            widget = self.layout().itemAt(i).widget()
            if widget and widget.objectName() == "save_buttons":
                self.layout().removeWidget(widget)
                widget.deleteLater()
        
        save_buttons = QWidget()
        save_buttons.setObjectName("save_buttons")
        button_layout = QHBoxLayout(save_buttons)
        
        btn_open_file = QPushButton("Fájl megnyitása")
        btn_open_file.setObjectName("success")
        btn_open_file.clicked.connect(lambda: open_with_application(save_path))
        
        btn_open_folder = QPushButton("Mappa megnyitása")
        btn_open_folder.setObjectName("success")
        btn_open_folder.clicked.connect(lambda: open_with_application(os.path.dirname(save_path)))
        
        button_layout.addWidget(btn_open_file)
        button_layout.addWidget(btn_open_folder)
        button_layout.addStretch()
        
        self.layout().insertWidget(self.layout().count() - 1, save_buttons)

if __name__ == "__main__":
    
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = FileSearchApp()
    window.show()
    sys.exit(app.exec_())